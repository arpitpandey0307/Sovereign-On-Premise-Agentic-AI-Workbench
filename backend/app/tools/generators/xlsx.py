"""XLSX generation.

Same contract as the Word generator: validated JSON in, a real workbook out,
no model anywhere near the bytes.

One rule worth stating because it is easy to get wrong: nothing written here
is ever a formula. A model that emits ``=SUM(B2:B9)`` is asking Excel to do
arithmetic nobody has checked, and a spreadsheet that computes a different
number when opened is worse than one that is plainly wrong. Values only --
the arithmetic belongs in ``python.execute``, where the result is captured.
"""

from __future__ import annotations

from typing import ClassVar

from app.artifacts.content import WorkbookContent
from app.artifacts.store import artifact_store
from app.tools.base import ToolContext, ToolResult

MAX_COLUMN_WIDTH = 60


class XlsxGenerateTool:
    name = "xlsx.generate"
    description = (
        "Build an Excel workbook from structured content: one or more sheets, "
        "each with columns and rows. Write computed values, never formulas -- "
        "use python.execute to calculate, then put the results here."
    )
    risk_level = "low"
    requires_approval = False
    input_schema: ClassVar[dict] = {
        "type": "object",
        "properties": {
            "title": {"type": "string"},
            "sheets": {"type": "array"},
            "filename": {"type": "string"},
        },
        "required": ["title", "sheets"],
    }

    def execute(self, args: dict, context: ToolContext) -> ToolResult:
        try:
            content = WorkbookContent.model_validate(
                {"title": args["title"], "sheets": args["sheets"]}
            )
        except Exception as exc:
            return ToolResult.failed(f"Content did not validate: {exc}")

        filename = args.get("filename") or "workbook.xlsx"
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"

        record = artifact_store.save(
            task_id=context.task_id,
            artifact_type="xlsx",
            filename=filename,
            payload=build_workbook(content),
        )
        return ToolResult(
            ok=True,
            data={
                "artifact_id": str(record.id),
                "filename": filename,
                "size_bytes": record.size_bytes,
                "sheets": [sheet.name for sheet in content.sheets],
            },
            detail=f"generated {filename}",
        )


def build_workbook(content: WorkbookContent) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_content in content.sheets:
        sheet = workbook.create_sheet(title=sheet_content.name[:31])
        sheet.append(sheet_content.columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in sheet_content.rows:
            sheet.append([_literal(value) for value in row])

        sheet.freeze_panes = "A2"
        _fit_columns(sheet, get_column_letter)

    workbook.properties.title = content.title
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _literal(value):
    """Never let a value be interpreted as a formula.

    openpyxl writes a leading "=" straight through, so a model emitting
    "=SUM(...)" -- or a document quoting one -- would become live arithmetic
    in the deliverable. Prefixing with an apostrophe forces Excel to treat it
    as text, which keeps what the agent computed and what the reader sees the
    same number.
    """
    if isinstance(value, str) and value.startswith("="):
        return f"'{value}"
    return value


def _fit_columns(sheet, get_column_letter) -> None:
    for index, column in enumerate(sheet.columns, start=1):
        longest = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=0,
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + 2, 10), MAX_COLUMN_WIDTH
        )
